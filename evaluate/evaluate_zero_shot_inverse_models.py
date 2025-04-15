#!/usr/bin/env python
# coding: utf-8

# In[1]:


import sys
sys.path.append("./")

import os
import json
import time
import datetime

import openpyxl
from openpyxl import Workbook


import torch
import torchvision
from torch import nn
from torchvision.transforms import Compose
from torch.utils.data import RandomSampler, DataLoader
from torch.utils.data.dataloader import default_collate
from torch.utils.data.distributed import DistributedSampler
from torch.nn.parallel import DistributedDataParallel
from torch.utils.tensorboard import SummaryWriter


from tqdm import tqdm
from tqdm.notebook import tqdm

from iunets import iUNet
from dataset import FWIDataset
# from networks import iunet_network
from networks import forward_network, inverse_network, iunet_network, autoencoder

import utils.transforms as T
from utils.pytorch_ssim import *
import utils.utilities as utils
from utils.scheduler import WarmupMultiStepLR
from utils.config_utils import get_config_name, get_latent_dim


# In[2]:


import argparse
import torch

parser = argparse.ArgumentParser(description="OpenFWI arguments")

# Define command-line arguments
parser.add_argument("--step", type=int, default=0, help="Step value")
parser.add_argument("--file_size", type=int, default=500, help="File size")
parser.add_argument("--vis_suffix", action="store_true", help="Enable visualization suffix")
parser.add_argument("--device", default="cuda", choices=["cpu", "cuda"], help="Device (cpu or cuda)")

parser.add_argument("--k", type=int, default=1, help="Value for k")
parser.add_argument("--workers", type=int, default=4, help="Number of workers")
parser.add_argument("--lambda_g1v", type=int, default=1, help="Value for lambda_g1v")
parser.add_argument("--lambda_g2v", type=int, default=1, help="Value for lambda_g2v")
parser.add_argument("--batch_size", type=int, default=50, help="Batch size")
parser.add_argument("--mask_factor", type=float, default=0.0, help="Mask factor")
parser.add_argument("--sample_temporal", type=int, default=1, help="Temporal sampling value")
parser.add_argument("--distributed", action="store_true", help="Enable distributed computing")

parser.add_argument("--num_images", type=int, default=5, help="Number of images")

parser.add_argument("--model_type", default="IUnetInverseModel", help="Model type")
parser.add_argument("--base_path", default="/projects/ml4science/OpenFWI/Results/SupervisedExperiment/", help="Base path")
parser.add_argument("--ckpt", default="latest_checkpoint.pth", help="Model type")
parser.add_argument("--model_save_name", default="IUnetInverseModel", help="Model save name")
parser.add_argument("--unet_depth", type=int, default=2, help="UNet depth")
parser.add_argument("--unet_repeat_blocks", type=int, default=2, help="Number of repeated UNet blocks")

parser.add_argument("--latent_dim", type=int, default=70, help="Latent Dimension")
parser.add_argument("--skip", type=int, default=1, help="Skip Connections for UNet")
parser.add_argument("--cfg_path", default="./configs/", help="Cfg path")

# Parse the command-line arguments
args = parser.parse_args()


# Reinitialize variables without the 'args' prefix
for key, value in vars(args).items():
    exec(f"{key} = value")


# In[3]:


def log_transform(data):
    return torch.log10(1+torch.abs(data)) * torch.sign(data)

def tanh_transform(data):
    return torch.nn.functional.tanh(data)


device = torch.device(device)
# ckpt = "latest_checkpoint.pth"

datasets = ["flatvel-a", "flatvel-b",
          "curvevel-a", "curvevel-b",
          "flatfault-a", "flatfault-b",
          "curvefault-a", "curvefault-b",
          "style-a", "style-b"]

model_names = ['FlatVel-A', 'FlatVel-B',
          'CurveVel-A', 'CurveVel-B',
         'FlatFault-A', 'FlatFault-B',
         'CurveFault-A', 'CurveFault-B',
         'Style-A', 'Style-B']




# #list of datasets on which model is evaluated
# datasets = ["flatvel-a", "flatvel-b"]
# #list of the datasets the model has been trained on
# model_names = ["FlatVel-A", "FlatVel-B"] 

model_paths = []
for model_name in model_names:
    path_ = os.path.join(model_name, model_save_name, "fcn_l1loss_ffb")
    model_paths.append(path_)
    
architecture_params = {"UNetInverseModel_17M":{"unet_depth": 2, "unet_repeat_blocks": 1}, 
                       "UNetInverseModel_33M":{"unet_depth": 2, "unet_repeat_blocks": 2},
                       "default":{"unet_depth": unet_depth, "unet_repeat_blocks": unet_repeat_blocks}
                      }


# In[4]:


criterions = {
    'MAE': lambda x, y: torch.mean(torch.abs(x - y)),
    'MSE': lambda x, y: torch.mean((x - y) ** 2)
}

def get_dataset_path(dataset):
    arr = dataset.split("-")
    base_path = f"./train_test_splits/"
    
    train_path = os.path.join(base_path, f"{arr[0]}_{arr[1]}_train.txt")
    val_path = os.path.join(base_path, f"{arr[0]}_{arr[1]}_val.txt")
    
    return train_path, val_path

def get_transforms(dataset, return_ctx=False):
    f = open('./dataset_config.json')
    ctx = json.load(f)[dataset]

    transform_data = T.Normalize(ctx['data_mean'], ctx['data_std'])
    transform_label = T.MinMaxNormalize(ctx['label_min'], ctx['label_max'])
    if return_ctx:
        return  transform_data, transform_label, ctx
    return  transform_data, transform_label


def get_dataloader(test_dataset, train_dataset=None):
    if train_dataset is None:
        train_dataset = test_dataset
    
    transform_data, transform_label, ctx = get_transforms(train_dataset, return_ctx=True)

    train_anno, val_anno = get_dataset_path(test_dataset)
        
    print(f'Loading {test_dataset} validation data')
    dataset_valid = FWIDataset(
        val_anno,
        preload=True,
        sample_ratio=sample_temporal,
        file_size=ctx['file_size'],
        transform_data=transform_data,
        transform_label=transform_label
    )
        
    valid_sampler = RandomSampler(dataset_valid)

    dataloader_valid = DataLoader(
                                dataset_valid, batch_size=batch_size,
                                sampler=valid_sampler, num_workers=workers,
                                pin_memory=True, collate_fn=default_collate)
    
    print('Data loading over')
        
    return dataset_valid, dataloader_valid, transform_data, transform_label 

def evaluate(model, dataloader, transform_data, transform_label, k, criterions, device):   
    
    eval_metrics = {}
    
    vel_list, vel_pred_list= [], [] # store denormalized velocity predcition & gt in numpy 
    vel_norm_list, vel_pred_norm_list = [], [] # store normalized velocity prediction & gt in tensor

    ssim_loss = SSIM(window_size=11)
    ssim_value = 0
    with torch.no_grad():
        batch_idx = 0
        for _, amp, vel in dataloader:
            amp = amp.to(device)
            vel = vel.to(device)
            
            vel_pred = model(amp)
            
            ssim_value += ssim_loss(vel / 2 + 0.5, vel_pred / 2 + 0.5).item()
            
            vel_np = transform_label.inverse_transform(vel.detach().cpu().numpy())
            vel_list.append(torch.from_numpy(vel_np))
            vel_norm_list.append(vel.detach().cpu())
            
            vel_pred_np = transform_label.inverse_transform(vel_pred.detach().cpu().numpy())
            vel_pred_list.append(torch.from_numpy(vel_pred_np))
            vel_pred_norm_list.append(vel_pred.detach().cpu())
            
            batch_idx += 1

    vel, vel_pred = torch.cat(vel_list), torch.cat(vel_pred_list)
    vel_norm, vel_pred_norm = torch.cat(vel_norm_list), torch.cat(vel_pred_norm_list)

    for name, criterion in criterions.items():
        
        eval_metrics[f'Velocity_norm_{name}'] = criterion(vel_norm, vel_pred_norm).item()
        eval_metrics[f'Velocity_unnorm_{name}'] = criterion(vel, vel_pred).item()
        
        
    eval_metrics[f'Velocity_SSIM']  = ssim_value/len(dataloader) 
#     ssim_loss = SSIM(window_size=11)
#     eval_metrics[f'Velocity_SSIM'] = ssim_loss(vel_norm / 2 + 0.5, vel_pred_norm / 2 + 0.5).item()
    return eval_metrics

def evaluate_iunet(model, dataloader, transform_data, transform_label, k, criterions, device):   
    print("Evaluating IUNET models.")
    eval_metrics = {}
    
    vel_list, vel_pred_list= [], [] # store denormalized velocity predcition & gt in numpy 
    vel_norm_list, vel_pred_norm_list = [], [] # store normalized velocity prediction & gt in tensor

    amp_list, amp_pred_list = [], []     # store denormalized waveform predcition & gt in numpy
    amp_norm_list, amp_pred_norm_list = [], []  # store normalized waveform predcition & gt in numpy

    ssim_loss = SSIM(window_size=11)
    ssim_vel = 0
    ssim_amp = 0
    with torch.no_grad():
        batch_idx = 0
        for _, amp, vel in dataloader:
            amp = amp.to(device)
            vel = vel.to(device)
            
            vel_pred = model.inverse(amp)
            amp_pred = model.forward(vel)
            
            ssim_vel += ssim_loss(vel / 2 + 0.5, vel_pred / 2 + 0.5).item()
            ssim_amp += ssim_loss(amp / 2 + 0.5, amp_pred / 2 + 0.5).item()
            
            vel_np = transform_label.inverse_transform(vel.detach().cpu().numpy())
            vel_list.append(torch.from_numpy(vel_np))
            vel_norm_list.append(vel.detach().cpu())
            
            vel_pred_np = transform_label.inverse_transform(vel_pred.detach().cpu().numpy())
            vel_pred_list.append(torch.from_numpy(vel_pred_np))
            vel_pred_norm_list.append(vel_pred.detach().cpu())
            
            
            amp_norm_list.append(amp.detach().cpu())
            amp_pred_norm_list.append(amp_pred.detach().cpu())
            
            amp_np = transform_data.inverse_transform(amp.detach().cpu().numpy())
            amp_pred_np = transform_data.inverse_transform(amp_pred.detach().cpu().numpy())
            
            amp_list.append(torch.from_numpy(amp_np))
            amp_pred_list.append(torch.from_numpy(amp_pred_np))
            
            batch_idx += 1

    vel, vel_pred = torch.cat(vel_list), torch.cat(vel_pred_list)
    vel_norm, vel_pred_norm = torch.cat(vel_norm_list), torch.cat(vel_pred_norm_list)
    
    amp, amp_pred = torch.cat(amp_list), torch.cat(amp_pred_list)
    amp_norm, amp_pred_norm = torch.cat(amp_norm_list), torch.cat(amp_pred_norm_list)
    
    amp_log, amp_pred_log = log_transform(amp), log_transform(amp_pred)
    amp_tanh, amp_pred_tanh = tanh_transform(amp), tanh_transform(amp_pred)
    
    for name, criterion in criterions.items():
        
        eval_metrics[f'Velocity_norm_{name}'] = criterion(vel_norm, vel_pred_norm).item()
        eval_metrics[f'Waveform_norm_{name}'] = criterion(amp_norm, amp_pred_norm).item()
        
        eval_metrics[f'Velocity_unnorm_{name}'] = criterion(vel, vel_pred).item()
        eval_metrics[f'Waveform_unnorm_{name}'] = criterion(amp, amp_pred).item()
        
        eval_metrics[f'Waveform_unnorm_log_{name}'] = criterion(amp_log, amp_pred_log).item()
        eval_metrics[f'Waveform_unnorm_tanh_{name}'] = criterion(amp_tanh, amp_pred_tanh).item()
        
        
#     ssim_loss = SSIM(window_size=11)
#     eval_metrics[f'Velocity_SSIM'] = ssim_loss(vel_norm / 2 + 0.5, vel_pred_norm / 2 + 0.5).item()
#     eval_metrics[f'Waveform_SSIM'] = ssim_loss(amp_norm, amp_pred_norm).item()
    
    eval_metrics[f'Velocity_SSIM']  = ssim_vel/len(dataloader) 
    eval_metrics[f'Waveform_SSIM']  = ssim_amp/len(dataloader) 
    
    return eval_metrics

def set_inverse_params(inverse_model_params):
        inverse_model_params.setdefault('IUnetInverseModel', {})
        inverse_model_params['IUnetInverseModel']['cfg_path'] = cfg_path
        inverse_model_params['IUnetInverseModel']['latent_dim'] = latent_dim
        
        inverse_model_params.setdefault('UNetInverseModel', {})
        inverse_model_params['UNetInverseModel']['cfg_path'] = cfg_path
        inverse_model_params['UNetInverseModel']['latent_dim'] = latent_dim
        inverse_model_params['UNetInverseModel']['unet_depth'] = architecture_params["default"]["unet_depth"]
        inverse_model_params['UNetInverseModel']['unet_repeat_blocks'] = architecture_params["default"]["unet_repeat_blocks"]
        inverse_model_params['UNetInverseModel']['skip'] = skip # skip true
        return inverse_model_params
    
def get_model(model_path, model_type):
#     try:
    print(model_path, model_type)
    inverse_model_params = inverse_network.inverse_params
    inverse_model_params = set_inverse_params(inverse_model_params)
    model = inverse_network.model_dict[model_type](**inverse_model_params[model_type]).to(device)
    checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
    model.load_state_dict(checkpoint['model'])
#     except:
#         print("Failed to load new model. Falling back to Legacy Code.")
#         inverse_model_params = inverse_network.inverse_params_legacy
#         inverse_model_params['unet_depth'] = unet_depth
#         inverse_model_params['unet_repeat_blocks'] = unet_repeat_blocks
#         model_type = model_type+"_Legacy"
#         model = inverse_network.model_dict[model_type](**inverse_model_params).to(device)
#         checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
#         model.load_state_dict(checkpoint['model'])

    model = model.to(device)
    model.eval()
    
    return model

def get_model_iunet_(amp_model, vel_model, latent_channels, model_type):
    if model_type == "IUNET":
        iunet_model = iUNet(in_channels=latent_channels, dim=2, architecture=(4,4,4,4))
        model = iunet_network.IUnetModel(amp_model, vel_model, iunet_model).to(device)
        print("IUnet model initialized.")
    elif model_type == "Decouple_IUnet":
        amp_iunet_model = iUNet(in_channels=latent_channels, dim=2, architecture=(4,4,4,4))
        vel_iunet_model = iUNet(in_channels=latent_channels, dim=2, architecture=(4,4,4,4))
        model = iunet_network.Decouple_IUnetModel(amp_model, vel_model, amp_iunet_model, vel_iunet_model).to(device)
        print("Decoupled IUnetModel model initialized.")
    else:
        print(f"Invalid Model: {model_type}")
    return model

def get_model_iunet(model_path, model_type):
    try:   
        print(model_path, model_type)
        amp_cfg_name = get_config_name(latent_dim, model_type="amplitude")
        amp_model = autoencoder.AutoEncoder(cfg_path, amp_cfg_name).to(device)

        # creating velocity cnn
        vel_cfg_name = get_config_name(latent_dim, model_type="velocity")
        vel_model = autoencoder.AutoEncoder(cfg_path, vel_cfg_name).to(device)

        latent_channels = get_latent_dim(cfg_path, amp_cfg_name)
        model = get_model_iunet_(amp_model, vel_model, latent_channels, model_type)
        checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
        model.load_state_dict(checkpoint['model'])
        
    except:
        print("Failed to load new model. Falling back to Legacy Code.")
        amp_input_channel = 5
        amp_encoder_channel = [8, 16, 32, 64, 128]
        amp_decoder_channel = [128, 64, 32, 16, 5]
        amp_model = iunet_network.AmpAutoEncoder(amp_input_channel, amp_encoder_channel, amp_decoder_channel).to(device)

        # creating velocity cnn
        vel_input_channel = 1
        vel_encoder_channel = [8, 16, 32, 64, 128]
        vel_decoder_channel = [128, 64, 32, 16, 1]
        vel_model = iunet_network.VelAutoEncoder(vel_input_channel, vel_encoder_channel, vel_decoder_channel).to(device)

        latent_channels = 128
        model = get_model_iunet_(amp_model, vel_model, latent_channels, model_type)
        checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
        model.load_state_dict(checkpoint['model'])

    model = model.to(device)
    model.eval()
    
    return model



def generate_eval_matrix(model_names, model_paths, datasets):
    
    metrics = {}

    for i in range(len(model_names)):
        model_name = model_names[i]
        model_path = os.path.join(base_path, model_paths[i], ckpt)
        if not os.path.exists(model_path):
            print(f"The path does not exist: {model_path}")
            continue
        if model_type == "IUNET":
            model = get_model_iunet(model_path, model_type)
        else:
            model = get_model(model_path, model_type)
        
        vis_path = os.path.join(base_path, model_paths[i], 'Zero_Shot_Generalization')
        if not os.path.exists(vis_path):
            os.makedirs(vis_path)
        
        # We need the transform_data/transform_label for the trained model.
        model_train_dataset = datasets[i]
        
        model_metrics = {}
        for dataset in datasets:
            dataset_val, dataloader_val,transform_data, transform_label= get_dataloader(test_dataset=dataset, train_dataset=model_train_dataset)
            print(f'------------ Evaluating {dataset} ------------')
            
            if model_type == "IUNET":
                eval_dict = evaluate_iunet(model, dataloader_val, transform_data, transform_label, k, criterions, device)
            else:
                eval_dict = evaluate(model, dataloader_val, transform_data, transform_label, k, criterions, device)
            utils.plot_images(num_images, dataset_val, model, dataset, vis_path, device, transform_data, transform_label, plot=False, save_key="dataset")
            
            for key in eval_dict.keys():
                model_metrics.setdefault(key, []).append(eval_dict[key])
                
        metrics[model_name] = model_metrics
    return metrics


def write_metrics(metrics_dict, filename):
    
    workbook = Workbook()
    workbook.remove(workbook.active)
    metrics_list = list(metrics_dict.values())[0].keys()
    
    for metric in metrics_list:
        workbook.create_sheet(title=metric)
        sheet = workbook[metric]
        
        for idx, dataset in enumerate(datasets, start=1):
            cell = sheet.cell(row=1, column=idx+1, value=dataset)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
        
        for idx, model_name in enumerate(metrics_dict.keys(), start=1):
            cell = sheet.cell(row=idx+1, column=1, value=model_name)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            metric_values = metrics_dict[model_name][metric]
            
            for col, val in enumerate(metric_values, start=1):
                cell = sheet.cell(row=idx+1, column=col+1, value=val)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    workbook.save(filename)


# In[5]:


eval_matrix = generate_eval_matrix(model_names, model_paths, datasets)


# In[16]:


write_metrics(eval_matrix, f'eval_metric{model_save_name}.xlsx')


# In[ ]:




