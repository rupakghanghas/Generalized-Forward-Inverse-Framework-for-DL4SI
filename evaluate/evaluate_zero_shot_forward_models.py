#!/usr/bin/env python
# coding: utf-8

# In[1]:


import sys
sys.path.append("./")

import os
import json
import time
import datetime

import openpyxl
from openpyxl import Workbook


import torch
import torchvision
from torch import nn
from torchvision.transforms import Compose
from torch.utils.data import RandomSampler, DataLoader
from torch.utils.data.dataloader import default_collate
from torch.utils.data.distributed import DistributedSampler
from torch.nn.parallel import DistributedDataParallel
from torch.utils.tensorboard import SummaryWriter


from tqdm.auto import tqdm
#from tqdm.notebook import tqdm
# from ipywidgets import FloatProgress

from iunets import iUNet
from dataset import FWIDataset
# from networks import iunet_network
from networks import forward_network, inverse_network, iunet_network, autoencoder

import utils.transforms as T
from utils.pytorch_ssim import *
import utils.utilities as utils
from utils.scheduler import WarmupMultiStepLR
from utils.config_utils import get_config_name, get_latent_dim


# In[2]:


import argparse
import torch

parser = argparse.ArgumentParser(description="OpenFWI arguments")

# Define command-line arguments
parser.add_argument("--step", type=int, default=0, help="Step value")
parser.add_argument("--file_size", type=int, default=500, help="File size")
parser.add_argument("--vis_suffix", action="store_true", help="Enable visualization suffix")
parser.add_argument("--device", default="cuda", choices=["cpu", "cuda"], help="Device (cpu or cuda)")

parser.add_argument("--k", type=int, default=1, help="Value for k")
parser.add_argument("--workers", type=int, default=4, help="Number of workers")
parser.add_argument("--lambda_g1v", type=int, default=1, help="Value for lambda_g1v")
parser.add_argument("--lambda_g2v", type=int, default=1, help="Value for lambda_g2v")
parser.add_argument("--batch_size", type=int, default=50, help="Batch size")
parser.add_argument("--mask_factor", type=float, default=0.0, help="Mask factor")
parser.add_argument("--sample_temporal", type=int, default=1, help="Temporal sampling value")
parser.add_argument("--distributed", action="store_true", help="Enable distributed computing")

parser.add_argument("--num_images", type=int, default=5, help="Number of images")

parser.add_argument("--model_type", default="IUnetForwardModel", help="Model type")
parser.add_argument("--base_path", default="/projects/ml4science/OpenFWI/Results/SupervisedExperiment/", help="Base path")
parser.add_argument("--model_save_name", default="IUnetForwardModel", help="Model save name")
parser.add_argument("--unet_depth", type=int, default=2, help="UNet depth")
parser.add_argument("--unet_repeat_blocks", type=int, default=2, help="Number of repeated UNet blocks")
parser.add_argument("--latent_dim", type=int, default=70, help="Latent Dimension")
parser.add_argument("--skip", type=int, default=1, help="Skip Connections for UNet")
parser.add_argument("--cfg_path", default="./configs/", help="Cfg path")



# Parse the command-line arguments
args = parser.parse_args()


# In[3]:


# Reinitialize variables without the 'args' prefix
for key, value in vars(args).items():
    exec(f"{key} = value")


# In[4]:


def log_transform(data):
    return torch.log10(1+torch.abs(data)) * torch.sign(data)

def tanh_transform(data):
    return torch.nn.functional.tanh(data)

datasets = ["flatvel-a", "flatvel-b",
          "curvevel-a", "curvevel-b",
          "flatfault-a", "flatfault-b",
          "curvefault-a", "curvefault-b",
          "style-a", "style-b"]

model_names = ['FlatVel-A', 'FlatVel-B',
          'CurveVel-A', 'CurveVel-B',
         'FlatFault-A', 'FlatFault-B',
         'CurveFault-A', 'CurveFault-B',
         'Style-A', 'Style-B']


# #list of datasets on which model is evaluated
# datasets = ["flatvel-a", "flatvel-b"]
# #list of the datasets the model has been trained on
# model_names = ["FlatVel-A", "FlatVel-B"] 

model_paths = []
for model_name in model_names:
    path_ = os.path.join(model_name, model_save_name, "fcn_l1loss_ffb")
    model_paths.append(path_)

architecture_params = {"UNetForwardModel_17M":{"unet_depth": 2, "unet_repeat_blocks": 1}, 
                       "UNetForwardModel_33M":{"unet_depth": 2, "unet_repeat_blocks": 2},
                       "default":{"unet_depth": 2, "unet_repeat_blocks": 2}
                      }
    
# model_paths = ["CurveVel-A/Joint/Cycle_Loss_0_Mask_Factor_80/lambda_amp_10_lambda_vel_1/fcn_l1loss_ffb",
#               "CurveVel-A/Joint/Cycle_Loss_0_Mask_Factor_80/lambda_amp_10_lambda_vel_1/fcn_l1loss_ffb"]


# In[5]:


criterions = {
    'MAE': lambda x, y: torch.mean(torch.abs(x - y)),
    'MSE': lambda x, y: torch.mean((x - y) ** 2)
}

def get_dataset_path(dataset):
    arr = dataset.split("-")
    base_path = f"./train_test_splits/"
    
    train_path = os.path.join(base_path, f"{arr[0]}_{arr[1]}_train.txt")
    val_path = os.path.join(base_path, f"{arr[0]}_{arr[1]}_val.txt")
    
    return train_path, val_path

def get_transforms(dataset, return_ctx=False):
    f = open('./dataset_config.json')
    ctx = json.load(f)[dataset]

    transform_data = T.Normalize(ctx['data_mean'], ctx['data_std'])
    transform_label = T.MinMaxNormalize(ctx['label_min'], ctx['label_max'])
    if return_ctx:
        return  transform_data, transform_label, ctx
    return  transform_data, transform_label


def get_dataloader(test_dataset, train_dataset=None):
    if train_dataset is None:
        train_dataset = test_dataset
    
    transform_data, transform_label, ctx = get_transforms(train_dataset, return_ctx=True)

    train_anno, val_anno = get_dataset_path(test_dataset)
        
    print(f'Loading {test_dataset} validation data')
    dataset_valid = FWIDataset(
        val_anno,
        preload=True,
        sample_ratio=sample_temporal,
        file_size=ctx['file_size'],
        transform_data=transform_data,
        transform_label=transform_label
    )
        
    valid_sampler = RandomSampler(dataset_valid)

    dataloader_valid = DataLoader(
                                dataset_valid, batch_size=batch_size,
                                sampler=valid_sampler, num_workers=workers,
                                pin_memory=True, collate_fn=default_collate)
    
    print('Data loading over')
        
    return dataset_valid, dataloader_valid, transform_data, transform_label 


def evaluate(model, dataloader, transform_data, transform_label, k, criterions, device):   
    eval_metrics = {}
    
    amp_list, amp_pred_list= [], [] # store denormalized velocity predcition & gt in numpy 
    amp_norm_list,amp_pred_norm_list = [], [] # store normalized velocity prediction & gt in tensor

    ssim_loss = SSIM(window_size=11)
    ssim_value = 0
    with torch.no_grad():
        batch_idx = 0
        for _, amp, vel in dataloader:
            amp = amp.to(device)
            vel = vel.to(device)
#             print("Getting prediction from model")
            amp_pred = model(vel)
#             print("model prediction received")

            ssim_value += ssim_loss(amp / 2 + 0.5, amp_pred / 2 + 0.5).item()
    
            amp_np = transform_data.inverse_transform(amp.detach().cpu().numpy())
            amp_list.append(torch.from_numpy(amp_np))
            amp_norm_list.append(amp.detach().cpu())
            
            amp_pred_np = transform_data.inverse_transform(amp_pred.detach().cpu().numpy())
            amp_pred_list.append(torch.from_numpy(amp_pred_np))
            amp_pred_norm_list.append(amp_pred.detach().cpu())
            
            batch_idx += 1

    amp, amp_pred = torch.cat(amp_list), torch.cat(amp_pred_list)
    amp_norm, amp_pred_norm = torch.cat(amp_norm_list), torch.cat(amp_pred_norm_list)
    
    amp_log, amp_pred_log = log_transform(amp), log_transform(amp_pred)
    amp_tanh, amp_pred_tanh = tanh_transform(amp), tanh_transform(amp_pred)
    
    for name, criterion in criterions.items():
        
        eval_metrics[f'Waveform_norm_{name}'] = criterion(amp_norm, amp_pred_norm).item()
        eval_metrics[f'Waveform_unnorm_{name}'] = criterion(amp, amp_pred).item()
        
        eval_metrics[f'Waveform_unnorm_log_{name}'] = criterion(amp_log, amp_pred_log).item()
        eval_metrics[f'Waveform_unnorm_tanh_{name}'] = criterion(amp_tanh, amp_pred_tanh).item()
        
    eval_metrics[f'Waveform_SSIM']  = ssim_value/len(dataloader) 
#     eval_metrics[f'Waveform_SSIM'] = ssim_loss(amp_norm / 2 + 0.5, amp_pred_norm / 2 + 0.5).item()
    return eval_metrics

def set_forward_params(forward_model_params):
        forward_model_params.setdefault('IUnetForwardModel', {})
        forward_model_params['IUnetForwardModel']['cfg_path'] = cfg_path
        forward_model_params['IUnetForwardModel']['latent_dim'] = latent_dim
        
        forward_model_params.setdefault('UNetForwardModel', {})
        forward_model_params['UNetForwardModel']['cfg_path'] = cfg_path
        forward_model_params['UNetForwardModel']['latent_dim'] = latent_dim
        forward_model_params['UNetForwardModel']['unet_depth'] = architecture_params["default"]["unet_depth"]
        forward_model_params['UNetForwardModel']['unet_repeat_blocks'] = architecture_params["default"]["unet_repeat_blocks"]
        forward_model_params['UNetForwardModel']['skip'] = skip # skip true
        return forward_model_params

def get_model(model_path, model_type):
    try:
        print(model_path)
        forward_model_params = forward_network.forward_params
        forward_model_params = set_forward_params(forward_model_params)
        model = forward_network.model_dict[model_type](**forward_model_params[model_type]).to(device)
        checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
        model.load_state_dict(checkpoint['model'])
    except:
        print("Failed to load new model. Falling back to Legacy Code.")
        forward_model_params = forward_network.forward_params_legacy
        forward_model_params['unet_depth'] = unet_depth
        forward_model_params['unet_repeat_blocks'] = unet_repeat_blocks
        model_type = model_type+"_Legacy"
        model = forward_network.model_dict[model_type](**forward_model_params).to(device)
        checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
        model.load_state_dict(checkpoint['model'])

    model = model.to(device)
    model.eval()
    
    return model

def generate_eval_matrix(model_names, model_paths, datasets):
    
    metrics = {}

    for i in range(len(model_names)):
        model_name = model_names[i]
        print(f'------------ Outer Loop: Evaluating {model_name} ------------')
        model_path = os.path.join(base_path, model_paths[i], "latest_checkpoint.pth")
        if not os.path.exists(model_path):
            print(f"The path does not exist: {model_path}")
            continue
        model = get_model(model_path, model_type)
        
        vis_path = os.path.join(base_path, model_paths[i], 'Zero_Shot_Generalization')
        if not os.path.exists(vis_path):
            os.makedirs(vis_path)
        
        # We need the transform_data/transform_label for the trained model.
        model_train_dataset = datasets[i]
        
        model_metrics = {}
        for dataset in datasets:
            dataset_val, dataloader_val,transform_data, transform_label  = get_dataloader(test_dataset=dataset, train_dataset=model_train_dataset)
            print(f'------------ Inner Loop Evaluating {dataset} ------------')
            eval_dict = evaluate(model, dataloader_val, transform_data, transform_label, k, criterions, device)
            utils.plot_images(num_images, dataset_val, model, dataset, vis_path, device, transform_data, transform_label, plot=False, save_key="dataset")
            
            for key in eval_dict.keys():
                model_metrics.setdefault(key, []).append(eval_dict[key])
                
        metrics[model_name] = model_metrics
    return metrics

def write_metrics(metrics_dict, filename):
    
    workbook = Workbook()
    workbook.remove(workbook.active)
    metrics_list = list(metrics_dict.values())[0].keys()
    
    for metric in metrics_list:
        workbook.create_sheet(title=metric)
        sheet = workbook[metric]
        
        for idx, dataset in enumerate(datasets, start=1):
            cell = sheet.cell(row=1, column=idx+1, value=dataset)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
        
        for idx, model_name in enumerate(metrics_dict.keys(), start=1):
            cell = sheet.cell(row=idx+1, column=1, value=model_name)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            metric_values = metrics_dict[model_name][metric]
            
            for col, val in enumerate(metric_values, start=1):
                cell = sheet.cell(row=idx+1, column=col+1, value=val)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    workbook.save(filename)


# In[ ]:


eval_matrix = generate_eval_matrix(model_names, model_paths, datasets)


# In[ ]:


write_metrics(eval_matrix, f'eval_metric{model_save_name}.xlsx')


# In[ ]:




