#!/usr/bin/env python
# coding: utf-8

# In[1]:


import sys
sys.path.append("../")

import os
import json
import time
import datetime

import openpyxl
from openpyxl import Workbook


import torch
import torchvision
from torch import nn
from torchvision.transforms import Compose
from torch.utils.data import RandomSampler, DataLoader
from torch.utils.data.dataloader import default_collate
from torch.utils.data.distributed import DistributedSampler
from torch.nn.parallel import DistributedDataParallel
from torch.utils.tensorboard import SummaryWriter


from tqdm.auto import tqdm
#from tqdm.notebook import tqdm
from ipywidgets import FloatProgress

from iunets import iUNet
from dataset import FWIDataset
# from networks import iunet_network
from networks import forward_network, inverse_network, iunet_network

import utils.transforms as T
from utils.pytorch_ssim import *
import utils.utilities as utils
from utils.scheduler import WarmupMultiStepLR


# In[2]:


step = 0
file_size = 500
vis_suffix = False
device = torch.device("cuda")

k = 1
workers = 4
lambda_g1v = 1
lambda_g2v = 1
batch_size = 50
mask_factor = 0.0
sample_temporal = 1
distributed = False

num_images = 5

base_path = "/projects/ml4science/OpenFWI/Results/SupervisedExperiment/"

model_type = "IUnetForwardModel"
model_save_name = "IUnetForwardModel"



model_type = "UNetForwardModel"
model_save_name = "UNetForwardModel_17M"

# model_type = "UNetForwardModel"
# model_save_name = "UNetForwardModel_33M"


unet_depth = 2
unet_repeat_blocks = 1

datasets = ["flatvel-a", "flatvel-b",
          "curvevel-a", "curvevel-b",
          "flatfault-a", "flatfault-b",
          "curvefault-a", "curvefault-b"]

model_names = ['FlatVel-A', 'FlatVel-B',
          'CurveVel-A', 'CurveVel-B',
         'FlatFault-A', 'FlatFault-B',
         'CurveFault-A', 'CurveFault-B']

# #list of datasets on which model is evaluated
# datasets = ["flatvel-a", "flatvel-b"]
# #list of the datasets the model has been trained on
# model_names = ["FlatVel-A", "FlatVel-B"] 

model_paths = []
for model_name in model_names:
    path_ = os.path.join(model_name, model_save_name, "fcn_l1loss_ffb")
    model_paths.append(path_)
    



criterions = {
    'MAE': lambda x, y: torch.mean(torch.abs(x - y)),
    'MSE': lambda x, y: torch.mean((x - y) ** 2)
}

def get_dataset_path(dataset):
    arr = dataset.split("-")
    base_path = f"./train_test_splits/"
    
    train_path = os.path.join(base_path, f"{arr[0]}_{arr[1]}_train.txt")
    val_path = os.path.join(base_path, f"{arr[0]}_{arr[1]}_val.txt")
    
    return train_path, val_path

def get_dataloader(dataset):
        
    f = open('./dataset_config.json')
    ctx = json.load(f)[dataset]

    transform_data = T.Normalize(ctx['data_mean'], ctx['data_std'])
    transform_label = T.MinMaxNormalize(ctx['label_min'], ctx['label_max'])

    train_anno, val_anno = get_dataset_path(dataset)
        
    print(f'Loading {dataset} validation data')
    dataset_valid = FWIDataset(
        val_anno,
        preload=True,
        sample_ratio=sample_temporal,
        file_size=ctx['file_size'],
        transform_data=transform_data,
        transform_label=transform_label
    )
        
    valid_sampler = RandomSampler(dataset_valid)

    dataloader_valid = DataLoader(
                                dataset_valid, batch_size=batch_size,
                                sampler=valid_sampler, num_workers=workers,
                                pin_memory=True, collate_fn=default_collate)
    
    print('Data loading over')
        
    return dataset_valid, dataloader_valid, transform_data, transform_label 

def evaluate(model, dataloader, transform_data, transform_label, k, criterions, device):   
    
    eval_metrics = {}
    
    amp_list, amp_pred_list= [], [] # store denormalized velocity predcition & gt in numpy 
    amp_norm_list,amp_pred_norm_list = [], [] # store normalized velocity prediction & gt in tensor

    with torch.no_grad():
        batch_idx = 0
        for _, amp, vel in dataloader:
            amp = amp.to(device)
            vel = vel.to(device)
            
            amp_pred = model(vel)
            
            amp_np = transform_data.inverse_transform(amp.detach().cpu().numpy())
            amp_list.append(torch.from_numpy(amp_np))
            amp_norm_list.append(amp.detach().cpu())
            
            amp_pred_np = transform_data.inverse_transform(amp_pred.detach().cpu().numpy())
            amp_pred_list.append(torch.from_numpy(amp_pred_np))
            amp_pred_norm_list.append(amp_pred.detach().cpu())
            
            batch_idx += 1

    amp, amp_pred = torch.cat(amp_list), torch.cat(amp_pred_list)
    amp_norm, amp_pred_norm = torch.cat(amp_norm_list), torch.cat(amp_pred_norm_list)

    for name, criterion in criterions.items():
        
        eval_metrics[f'Waveform_norm_{name}'] = criterion(amp_norm, amp_pred_norm).item()
        eval_metrics[f'Waveform_unnorm_{name}'] = criterion(amp, amp_pred).item()
        
        
    ssim_loss = SSIM(window_size=11)
    eval_metrics[f'Waveform_SSIM'] = ssim_loss(amp_norm / 2 + 0.5, amp_pred_norm / 2 + 0.5).item()
    return eval_metrics

def get_model(model_path, model_type):
    forward_model_params = forward_network.forward_params
    forward_model_params['unet_depth'] = unet_depth
    forward_model_params['unet_repeat_blocks'] = unet_repeat_blocks
    model = forward_network.model_dict[model_type](**forward_model_params).to(device)

    checkpoint = torch.load(model_path, map_location=torch.device('cpu'))
    model.load_state_dict(checkpoint['model'])
    model = model.to(device)
    model.eval()
    return model


def generate_eval_matrix(model_names, model_paths, datasets):
    
    metrics = {}

    for i in range(len(model_names)):
        model_name = model_names[i]
        print(f'------------ Outer Loop: Evaluating {model_name} ------------')
        model_path = os.path.join(base_path, model_paths[i], "latest_checkpoint.pth")
        model = get_model(model_path, model_type)
        
        vis_path = os.path.join(base_path, model_paths[i], 'Zero_Shot_Generalization')
        if not os.path.exists(vis_path):
            os.makedirs(vis_path)
        
        model_metrics = {}
        for dataset in datasets:
            dataset_val, dataloader_val, transform_data, transform_label = get_dataloader(dataset)
            print(f'------------ Inner Loop Evaluating {dataset} ------------')
            eval_dict = evaluate(model, dataloader_val, transform_data, transform_label, k, criterions, device)
            utils.plot_images(num_images, dataset_val, model, dataset, vis_path, device, transform_data, transform_label, plot=False, save_key="dataset")
            
            for key in eval_dict.keys():
                model_metrics.setdefault(key, []).append(eval_dict[key])
                
        metrics[model_name] = model_metrics
    return metrics


def write_metrics(metrics_dict, filename):
    
    workbook = Workbook()
    workbook.remove(workbook.active)
    metrics_list = list(metrics_dict.values())[0].keys()
    
    for metric in metrics_list:
        workbook.create_sheet(title=metric)
        sheet = workbook[metric]
        
        for idx, dataset in enumerate(datasets, start=1):
            cell = sheet.cell(row=1, column=idx+1, value=dataset)
            cell.font = openpyxl.styles.Font(bold=True)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
        
        for idx, model_name in enumerate(metrics_dict.keys(), start=1):
            cell = sheet.cell(row=idx+1, column=1, value=model_name)
            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            metric_values = metrics_dict[model_name][metric]
            
            for col, val in enumerate(metric_values, start=1):
                cell = sheet.cell(row=idx+1, column=col+1, value=val)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    workbook.save(filename)


# In[ ]:


eval_matrix = generate_eval_matrix(model_names, model_paths, datasets)


# In[ ]:


write_metrics(eval_matrix, 'eval_matrix_unet_17m_forward_model.xlsx')


# In[ ]:




